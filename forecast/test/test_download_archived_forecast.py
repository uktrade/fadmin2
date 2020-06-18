import io

from django.contrib.auth.models import Permission
from django.test import TestCase
from django.urls import reverse

from openpyxl import load_workbook

from end_of_month.test.test_utils import SetFullYearArchive

from chartofaccountDIT.test.factories import (
    NaturalCodeFactory,
    ProgrammeCodeFactory,
    ProjectCodeFactory,
)

from core.models import FinancialYear
from core.myutils import get_current_financial_year
from core.test.test_base import RequestFactoryBase

from costcentre.test.factories import (
    CostCentreFactory,
    DepartmentalGroupFactory,
    DirectorateFactory,
)

from forecast.models import (
    FinancialCode,
    FinancialPeriod,
    ForecastMonthlyFigure,
)
from forecast.permission_shortcuts import assign_perm
from forecast.test.test_utils import create_budget
from forecast.views.view_forecast.export_forecast_data import (
    export_forecast_data_cost_centre,
    export_forecast_data_directorate,
    export_forecast_data_dit,
    export_forecast_data_group,
)

BUDGET_CELL = "Y2"
TOTAL_CELL = "AO2"
GROUP_HEADING_CELL = "A1"
GROUP_CODE_CELL = "B2"

class DownloadForecastHierarchyTest(TestCase, RequestFactoryBase):
    def setUp(self):
        RequestFactoryBase.__init__(self)

        # Assign forecast view permission
        can_view_forecasts = Permission.objects.get(
            codename='can_view_forecasts'
        )
        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()
        self.archive = SetFullYearArchive()

    def test_dit_download(self):
        test_period = 2
        dit_url = self.factory_get(
            reverse("export_forecast_data_dit",
                    kwargs={"period": test_period}),
            export_forecast_data_dit,
            period=test_period,
        )

        self.assertEqual(dit_url.status_code, 200)

        file = io.BytesIO(dit_url.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        year_total = 0
        for col in range(26, 41):
            year_total += ws.cell(column=col, row=2,).value

        print(year_total)
        # for col in range(1, 50):
        #     print(f'{col}: {ws.cell(column=col, row=1,).coordinate} '
        #           f'{ws.cell(column=col, row=1,).value} {ws.cell(column=col, row=2,).value}')
        # Check group
        self.assertEqual(ws[GROUP_HEADING_CELL].value, "Group name")
        self.assertEqual(ws[GROUP_CODE_CELL].value, self.archive.group_code)

        self.assertEqual(ws[BUDGET_CELL].value,
                         self.archive.archived_budget[test_period] / 100)

        self.assertEqual(year_total,
                         self.archive.archived_forecast[test_period] / 100)

    def test_group_download(self):
        test_period = 2
        response = self.factory_get(
            reverse(
                "export_forecast_data_group",
                kwargs={
                    'group_code': self.archive.group_code,
                    'period': test_period,
                },
            ),
            export_forecast_data_group,
            group_code=self.archive.group_code,
            period=test_period,
        )

        self.assertEqual(response.status_code, 200)

        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        # Check group
        assert ws[GROUP_HEADING_CELL].value == "Group name"
        assert ws[GROUP_CODE_CELL].value == self.archive.group_code

    def test_directorate_download(self):
        test_period = 3
        response = self.factory_get(
            reverse(
                "export_forecast_data_directorate",
                kwargs={
                    'directorate_code': self.archive.directorate_code,
                    'period': test_period,
                },
            ),
            export_forecast_data_directorate,
            directorate_code=self.archive.directorate_code,
            period=test_period,
        )

        self.assertEqual(response.status_code, 200)

        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file, read_only=True)
        ws = wb.active
        # Check group
        assert ws[GROUP_HEADING_CELL].value == "Group name"
        assert ws[GROUP_CODE_CELL].value == self.archive.group_code


    def test_cost_centre_download(self):
        test_period = 4
        response = self.factory_get(
            reverse(
                "export_forecast_data_cost_centre",
                kwargs={
                    'cost_centre': self.archive.cost_centre_code,
                    'period': test_period,
                },
            ),
            export_forecast_data_cost_centre,
            cost_centre=self.archive.cost_centre_code,
            period=test_period,
        )
        self.assertEqual(response.status_code, 200)

        file = io.BytesIO(response.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        # Check group
        assert ws[GROUP_HEADING_CELL].value == "Group name"
        assert ws[GROUP_CODE_CELL].value == self.archive.group_code

