import io

from django.contrib.auth.models import Permission
from django.test import TestCase
from django.urls import reverse

from openpyxl import load_workbook

from chartofaccountDIT.test.factories import (
    NaturalCodeFactory,
    ProgrammeCodeFactory,
    ProjectCodeFactory,
)

from core.models import FinancialYear
from core.myutils import get_current_financial_year
from core.test.test_base import RequestFactoryBase

from costcentre.test.factories import (
    CostCentreFactory,
    DepartmentalGroupFactory,
    DirectorateFactory,
)

from forecast.models import (
    FinancialCode,
    FinancialPeriod,
    ForecastMonthlyFigure,
)
from forecast.permission_shortcuts import assign_perm
from forecast.test.test_utils import create_budget
from forecast.views.view_forecast.export_forecast_data import (
    export_edit_forecast_data,
)


class DownloadEditForecastTest(TestCase, RequestFactoryBase):
    def setUp(self):
        RequestFactoryBase.__init__(self)

        self.group_name = "Test Group"
        self.group_code = "TestGG"
        self.directorate_name = "Test Directorate"
        self.directorate_code = "TestDD"
        self.cost_centre_code = 109076

        can_view_forecasts = Permission.objects.get(
            codename='can_view_forecasts'
        )

        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()

        self.group = DepartmentalGroupFactory(
            group_code=self.group_code,
            group_name=self.group_name,
        )
        self.directorate = DirectorateFactory(
            directorate_code=self.directorate_code,
            directorate_name=self.directorate_name,
            group=self.group,
        )
        self.cost_centre = CostCentreFactory(
            directorate=self.directorate,
            cost_centre_code=self.cost_centre_code,
        )
        current_year = get_current_financial_year()
        self.amount_apr = -9876543
        self.programme_obj = ProgrammeCodeFactory()
        self.nac_obj = NaturalCodeFactory()
        self.project_obj = ProjectCodeFactory()
        year_obj = FinancialYear.objects.get(financial_year=current_year)

        apr_period = FinancialPeriod.objects.get(financial_period_code=1)
        apr_period.actual_loaded = True
        apr_period.save()

        # If you use the MonthlyFigureFactory the test fails.
        # I cannot work out why, it may be due to using a random year....
        financial_code_obj = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=self.cost_centre,
            natural_account_code=self.nac_obj,
            project_code=self.project_obj
        )
        financial_code_obj.save
        apr_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(
                financial_period_code=1
            ),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr
        )
        apr_figure.save
        self.amount_may = 1234567
        may_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(
                financial_period_code=2,
            ),
            amount=self.amount_may,
            financial_code=financial_code_obj,
            financial_year=year_obj
        )
        may_figure.save

        project_obj1 = ProjectCodeFactory.create(project_code='123456')

        financial_code_obj1 = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=self.cost_centre,
            natural_account_code=self.nac_obj,
            project_code=project_obj1
        )
        financial_code_obj1.save

        # Assign forecast view permission
        can_view_forecasts = Permission.objects.get(
            codename='can_view_forecasts'
        )
        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()

        self.budget = create_budget(financial_code_obj, year_obj)
        self.year_total = (self.amount_apr + self.amount_may)/100
        self.underspend_total = self.budget/100 - self.year_total
        self.spend_to_date_total = self.amount_apr/100

    def test_user_can_download_forecast(self):
        """
        User can access download URL if they have cost centre editing permission
        """

        assign_perm("change_costcentre", self.test_user, self.cost_centre)
        self.cost_centre_code = self.cost_centre

        download_forecast_url = self.factory_get(
            reverse(
                "export_edit_forecast_data_cost_centre",
                kwargs={
                    'cost_centre': self.cost_centre_code
                },
            ),
            export_edit_forecast_data,
            cost_centre=self.cost_centre.cost_centre_code,
        )
        self.assertEqual(download_forecast_url.status_code, 200)

        file = io.BytesIO(download_forecast_url.content)
        wb = load_workbook(filename=file, )
        ws = wb.active
        print(f'ws["I1"].value == {ws["I1"].value}')
        print(f'ws["I2"].value == {ws["I2"].value}')
        print(f'ws["I3"].value == {ws["I3"].value}')
        print(f'ws["I4"].value == {ws["I4"].value}')

        print(f'ws["J1"].value == {ws["J1"].value}')
        print(f'ws["J2"].value == {ws["J2"].value}')
        print(f'ws["J3"].value == {ws["J3"].value}')
        print(f'ws["J4"].value == {ws["J4"].value}')

        print(f'ws["X1"].value == {ws["X1"].value}')
        print(f'ws["X2"].value == {ws["X2"].value}')
        print(f'ws["X3"].value == {ws["X3"].value}')
        print(f'ws["X4"].value == {ws["X4"].value}')

        print(f'ws["Y1"].value == {ws["Y1"].value}')
        print(f'ws["Y2"].value == {ws["Y2"].value}')
        print(f'ws["Y3"].value == {ws["Y3"].value}')
        print(f'ws["Y4"].value == {ws["Y4"].value}')

        print(f'ws["Z1"].value == {ws["Z1"].value}')
        print(f'ws["Z2"].value == {ws["Z2"].value}')
        print(f'ws["Z3"].value == {ws["Z3"].value}')
        print(f'ws["Z4"].value == {ws["Z4"].value}')



        print(f'self.budget {self.budget}')
        print(f'self.amount_apr {self.amount_apr}')
        print(f'self.amount_may {self.amount_may}')
        print(f'self.year_total {self.year_total}')
        print(f'self.underspend_total {self.underspend_total}')
        print(f'self.spend_to_date_total {self.spend_to_date_total}')
        assert ws["C1"].value == "Natural Account code"
        assert ws["C2"].value == self.nac_obj.natural_account_code
        assert ws["I1"].value == 'Apr'
        assert ws.max_row == 4
        ws["I1"].value == 'Apr'
        ws["I2"].value == 0
        ws["I3"].value == -98765.43
        ws["I4"].value == '= SUM(I2:I3)'
        ws["J1"].value == 'May'
        ws["J2"].value == 0
        ws["J3"].value == 12345.67
        ws["J4"].value == '= SUM(J2:J3)'
        ws["X1"].value == 'Forecast outturn'
        ws["X2"].value == ''= SUM(I2:W2)''
        ws["X3"].value == '= SUM(I3:W3)''
        ws["X4"].value == ''= SUM(X2:X3)''
        ws["Y1"].value == Variance - overspend / underspend
        ws["Y2"].value == = (H2 - X2)
        ws["Y3"].value == = (H3 - X3)
        ws["Y4"].value == = SUM(Y2:Y3)
        ws["Z1"].value == Year
        to
        Date
        Actuals
        ws["Z2"].value == = SUM(I2:I2)
        ws["Z3"].value == = SUM(I3:I3)
        ws["Z4"].value == = SUM(Z2:Z3)

        wb.save("test.xlsx")
    def test_user_cannot_download_forecast(self):
        """
        User can't access download URL if they have no editing permission
        """
        self.cost_centre_code = self.cost_centre

        download_forecast_url = self.factory_get(
            reverse(
                "export_edit_forecast_data_cost_centre",
                kwargs={
                    'cost_centre': self.cost_centre_code
                },
            ),
            export_edit_forecast_data,
            cost_centre=self.cost_centre.cost_centre_code,
        )
        self.assertEqual(download_forecast_url.status_code, 302)

    def test_user_cannot_download_wrong_forecast(self):
        """
            User can't access download URL if they do not
            have editing permission for certain Cost Centre
        """
        # Assigns user to one cost centre
        assign_perm("change_costcentre", self.test_user, self.cost_centre)

        # Changes cost_centre_code to one that user can view but NOT edit
        self.test_cost_centre = 888332
        self.cost_centre_code = self.test_cost_centre
        self.cost_centre = CostCentreFactory.create(
            cost_centre_code=self.cost_centre_code
        )

        download_forecast_url = self.factory_get(
            reverse(
                "export_edit_forecast_data_cost_centre",
                kwargs={
                    'cost_centre': self.cost_centre_code
                },
            ),
            export_edit_forecast_data,
            cost_centre=self.cost_centre.cost_centre_code,
        )
        self.assertEqual(download_forecast_url.status_code, 302)

    def test_download(self):
        assign_perm("change_costcentre", self.test_user, self.cost_centre)
        self.cost_centre_code = self.cost_centre

        download_forecast_url = self.factory_get(
            reverse(
                "export_edit_forecast_data_cost_centre",
                kwargs={
                    'cost_centre': self.cost_centre_code
                },
            ),
            export_edit_forecast_data,
            cost_centre=self.cost_centre.cost_centre_code,
        )
        self.assertEqual(download_forecast_url.status_code, 200)
        file = io.BytesIO(download_forecast_url.content)
        wb = load_workbook(filename=file)
        ws = wb.active
        assert ws["C1"].value == "Natural Account code"
        assert ws["C2"].value == self.nac_obj.natural_account_code
        assert ws["I1"].value == 'Apr'
        assert ws.max_row == 4

        # Check that the variance over/under spend is correct

        # Check the existence of grand totals