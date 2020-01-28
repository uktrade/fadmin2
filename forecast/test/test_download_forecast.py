import io
import openpyxl

from bs4 import BeautifulSoup

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.exceptions import PermissionDenied
from django.test import TestCase
from django.urls import reverse

from chartofaccountDIT.test.factories import (
    Analysis1Factory,
    Analysis2Factory,
    ExpenditureCategoryFactory,
    NaturalCodeFactory,
    ProgrammeCodeFactory,
    ProjectCodeFactory,
)

from core.models import FinancialYear
from core.myutils import get_current_financial_year
from core.test.test_base import RequestFactoryBase

from costcentre.test.factories import (
    CostCentreFactory,
    DepartmentalGroupFactory,
    DirectorateFactory,
)

from forecast.models import (
    BudgetMonthlyFigure,
    FinancialCode,
    FinancialPeriod,
    ForecastMonthlyFigure,
)

from forecast.views.view_forecast.export_forecast_data import  (
    export_forecast_data_dit,
    export_forecast_data_cost_centre,
    export_forecast_data_directorate,
    export_forecast_data_group,
)

from forecast.permission_shortcuts import assign_perm
from forecast.test.test_utils import create_budget
from forecast.views.edit_forecast import (
    AddRowView,
    ChooseCostCentreView,
    EditForecastView,
)
from forecast.views.view_forecast.expenditure_details import (
    CostCentreExpenditureDetailsView,
    DITExpenditureDetailsView,
    DirectorateExpenditureDetailsView,
    GroupExpenditureDetailsView,
)
from forecast.views.view_forecast.forecast_summary import (
    CostCentreView,
    DITView,
    DirectorateView,
    GroupView,
)
from forecast.views.view_forecast.programme_details import (
    DITProgrammeDetailsView,
    DirectorateProgrammeDetailsView,
    GroupProgrammeDetailsView,
)

class ViewForecastHierarchyTest(TestCase, RequestFactoryBase):
    def setUp(self):
        RequestFactoryBase.__init__(self)

        self.group_name = "Test Group"
        self.group_code = "TestGG"
        self.directorate_name = "Test Directorate"
        self.directorate_code = "TestDD"
        self.cost_centre_code = 109076

        self.group = DepartmentalGroupFactory(
            group_code=self.group_code,
            group_name=self.group_name,
        )
        self.directorate = DirectorateFactory(
            directorate_code=self.directorate_code,
            directorate_name=self.directorate_name,
            group=self.group,
        )
        self.cost_centre = CostCentreFactory(
            directorate=self.directorate,
            cost_centre_code=self.cost_centre_code,
        )
        current_year = get_current_financial_year()
        self.amount_apr = -9876543
        self.programme_obj = ProgrammeCodeFactory()
        nac_obj = NaturalCodeFactory()
        self.project_obj = ProjectCodeFactory()
        year_obj = FinancialYear.objects.get(financial_year=current_year)

        apr_period = FinancialPeriod.objects.get(financial_period_code=1)
        apr_period.actual_loaded = True
        apr_period.save()

        # If you use the MonthlyFigureFactory the test fails.
        # I cannot work out why, it may be due to using a random year....
        financial_code_obj = FinancialCode.objects.create(
            programme=self.programme_obj,
            cost_centre=self.cost_centre,
            natural_account_code=nac_obj,
            project_code=self.project_obj
        )
        financial_code_obj.save
        apr_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(
                financial_period_code=1
            ),
            financial_code=financial_code_obj,
            financial_year=year_obj,
            amount=self.amount_apr
        )
        apr_figure.save
        self.amount_may = 1234567
        may_figure = ForecastMonthlyFigure.objects.create(
            financial_period=FinancialPeriod.objects.get(
                financial_period_code=2,
            ),
            amount=self.amount_may,
            financial_code=financial_code_obj,
            financial_year=year_obj
        )
        may_figure.save
        # Assign forecast view permission
        can_view_forecasts = Permission.objects.get(
            codename='can_view_forecasts'
        )
        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()

        self.budget = create_budget(financial_code_obj, year_obj)
        self.year_total = self.amount_apr + self.amount_may
        self.underspend_total = self.budget - self.amount_apr - self.amount_may
        self.spend_to_date_total = self.amount_apr

    def test_dit_download(self):
        response = self.factory_get(
            reverse("export_forecast_data_dit"),
            export_forecast_data_dit,
        )

        self.assertEqual(response.status_code, 200)

        import io
        file = io.BytesIO(response.content)
        #  file = open("book_1.xlsx","rb")
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.active
        # Check group is shown
        assert ws["A1"] == "Group name"
        assert ws["B2"] == self.group_code



    # def test_group_view(self):
    #     response = self.factory_get(
    #         reverse(
    #             "forecast_group",
    #             kwargs={
    #                 'group_code': self.group.group_code
    #             },
    #         ),
    #         GroupView,
    #         group_code=self.group.group_code,
    #     )
    #     self.assertEqual(response.status_code, 200)
    #
    #     # Check directorate is shown
    #     assert self.directorate_name in str(response.rendered_content)
    #
    # def test_directorate_view(self):
    #     response = self.factory_get(
    #         reverse(
    #             "forecast_directorate",
    #             kwargs={
    #                 'directorate_code': self.directorate.directorate_code
    #             },
    #         ),
    #         DirectorateView,
    #         directorate_code=self.directorate.directorate_code,
    #     )
    #     self.assertEqual(response.status_code, 200)
    #
    #     # Check cost centre is shown
    #     assert str(self.cost_centre_code) in str(response.rendered_content)
    #
    # def test_cost_centre_view(self):
    #     response = self.factory_get(
    #         reverse(
    #             "forecast_cost_centre",
    #             kwargs={
    #                 'cost_centre_code': self.cost_centre_code
    #             },
    #         ),
    #         CostCentreView,
    #         cost_centre_code=self.cost_centre.cost_centre_code,
    #     )
    #     self.assertEqual(response.status_code, 200)
    #
    #     # Check directorate is shown
    #     assert str(self.cost_centre_code) in str(response.rendered_content)
    #
    # def check_programme_table(self, table, prog_index=2):
    #     programme_rows = table.find_all("tr")
    #     first_prog_cols = programme_rows[1].find_all("td")
    #     assert first_prog_cols[prog_index].get_text() == \
    #         self.programme_obj.programme_description
    #     assert first_prog_cols[prog_index + 1].get_text() == \
    #         self.programme_obj.programme_code
    #
    #     last_programme_cols = programme_rows[-1].find_all("td")
    #     # Check the total for the year
    #     assert last_programme_cols[TOTAL_COLUMN].get_text() == \
    #         format_forecast_figure(self.year_total / 100)
    #     # Check the difference between budget and year total
    #     assert last_programme_cols[UNDERSPEND_COLUMN].get_text() == \
    #         format_forecast_figure(self.underspend_total / 100)
    #     # Check the spend to date
    #     assert last_programme_cols[SPEND_TO_DATE_COLUMN].get_text() == \
    #         format_forecast_figure(self.spend_to_date_total / 100)
    #
    # def check_expenditure_table(self, table):
    #     expenditure_rows = table.find_all("tr")
    #     first_expenditure_cols = expenditure_rows[1].find_all("td")
    #     assert (first_expenditure_cols[1].get_text() == '—')
    #     assert first_expenditure_cols[3].get_text() == format_forecast_figure(
    #         self.budget / 100
    #     )
    #
    #     last_expenditure_cols = expenditure_rows[-1].find_all("td")
    #     # Check the total for the year
    #     assert last_expenditure_cols[TOTAL_COLUMN].get_text() == \
    #         format_forecast_figure(self.year_total / 100)
    #     # Check the difference between budget and year total
    #     assert last_expenditure_cols[UNDERSPEND_COLUMN].get_text() == \
    #         format_forecast_figure(self.underspend_total / 100)
    #     # Check the spend to date
    #     assert last_expenditure_cols[SPEND_TO_DATE_COLUMN].get_text() == \
    #         format_forecast_figure(self.spend_to_date_total / 100)
    #
    # def check_project_table(self, table):
    #     project_rows = table.find_all("tr")
    #     first_project_cols = project_rows[1].find_all("td")
    #     assert first_project_cols[1].get_text() == self.project_obj.project_description
    #     assert first_project_cols[2].get_text() == self.project_obj.project_code
    #     assert first_project_cols[3].get_text() == format_forecast_figure(
    #         self.budget / 100
    #     )
    #
    #     last_project_cols = project_rows[-1].find_all("td")
    #     # Check the total for the year
    #     assert last_project_cols[TOTAL_COLUMN].get_text() == \
    #         format_forecast_figure(self.year_total / 100)
    #     # Check the difference between budget and year total
    #     assert last_project_cols[UNDERSPEND_COLUMN].get_text() == \
    #         format_forecast_figure(self.underspend_total / 100)
    #     # Check the spend to date
    #     assert last_project_cols[SPEND_TO_DATE_COLUMN].get_text() == \
    #         format_forecast_figure(self.spend_to_date_total / 100)
    #
    # def check_hierarchy_table(self, table, hierarchy_element):
    #     hierarchy_rows = table.find_all("tr")
    #     first_hierarchy_cols = hierarchy_rows[1].find_all("td")
    #     assert first_hierarchy_cols[1].get_text() == hierarchy_element
    #
    #     assert first_hierarchy_cols[3].get_text() == format_forecast_figure(
    #         self.budget / 100
    #     )
    #     assert first_hierarchy_cols[4].get_text() == format_forecast_figure(
    #         self.amount_apr / 100
    #     )
    #
    #     last_hierarchy_cols = hierarchy_rows[-1].find_all("td")
    #     # Check the total for the year
    #     assert last_hierarchy_cols[TOTAL_COLUMN].get_text() == \
    #         format_forecast_figure(self.year_total / 100)
    #     # Check the difference between budget and year total
    #     assert last_hierarchy_cols[UNDERSPEND_COLUMN].get_text() == \
    #         format_forecast_figure(self.underspend_total / 100)
    #     # Check the spend to date
    #     assert last_hierarchy_cols[SPEND_TO_DATE_COLUMN].get_text() == \
    #         format_forecast_figure(self.spend_to_date_total / 100)
    #
    # def check_negative_value_formatted(self, soup):
    #     negative_values = soup.find_all("span", class_="negative")
    #     assert len(negative_values) == 42
    #
    # def test_view_cost_centre_summary(self):
    #     resp = self.factory_get(
    #         reverse(
    #             "forecast_cost_centre",
    #             kwargs={
    #                 'cost_centre_code': self.cost_centre_code
    #             },
    #         ),
    #         CostCentreView,
    #         cost_centre_code=self.cost_centre_code,
    #     )
    #
    #     self.assertEqual(resp.status_code, 200)
    #     self.assertContains(resp, "govuk-table")
    #     soup = BeautifulSoup(resp.content, features="html.parser")
    #     # Check that there are 4 tables on the page
    #     tables = soup.find_all("table", class_="govuk-table")
    #     assert len(tables) == 4
    #
    #     # Check that the first table displays the cost centre code
    #
    #     # Check that all the subtotal hierachy_rows exist
    #     table_rows = soup.find_all("tr", class_="govuk-table__row")
    #     assert len(table_rows) == 18
    #
    #     self.check_negative_value_formatted(soup)
    #
    #     self.check_hierarchy_table(tables[HIERARCHY_TABLE_INDEX],
    #                                self.cost_centre.cost_centre_name)
    #     # Check that the second table displays the programme and the correct totals
    #     # The programme table in the cost centre does not show the 'View'
    #     # so the programme is displayed in a different column
    #     self.check_programme_table(tables[PROGRAMME_TABLE_INDEX], 1)
    #
    #     # Check that the third table displays the expenditure and the correct totals
    #     self.check_expenditure_table(tables[EXPENDITURE_TABLE_INDEX])
    #
    #     # Check that the second table displays the project and the correct totals
    #     self.check_project_table(tables[PROJECT_TABLE_INDEX])
    #
    # def test_view_directorate_summary(self):
    #     resp = self.factory_get(
    #         reverse(
    #             "forecast_directorate",
    #             kwargs={
    #                 'directorate_code': self.directorate.directorate_code
    #             },
    #         ),
    #         DirectorateView,
    #         directorate_code=self.directorate.directorate_code,
    #     )
    #
    #     self.assertEqual(resp.status_code, 200)
    #     self.assertContains(resp, "govuk-table")
    #     soup = BeautifulSoup(resp.content, features="html.parser")
    #
    #     # Check that there are 4 tables on the page
    #     tables = soup.find_all("table", class_="govuk-table")
    #     assert len(tables) == 4
    #
    #     # Check that the first table displays the cost centre code
    #
    #     # Check that all the subtotal hierachy_rows exist
    #     table_rows = soup.find_all("tr", class_="govuk-table__row")
    #     assert len(table_rows) == 18
    #
    #     self.check_negative_value_formatted(soup)
    #
    #     self.check_hierarchy_table(tables[HIERARCHY_TABLE_INDEX],
    #                                self.cost_centre.cost_centre_name)
    #     # Check that the second table displays the programme and the correct totals
    #     self.check_programme_table(tables[PROGRAMME_TABLE_INDEX])
    #
    #     # Check that the third table displays the expenditure and the correct totals
    #     self.check_expenditure_table(tables[EXPENDITURE_TABLE_INDEX])
    #
    #     # Check that the second table displays the project and the correct totals
    #     self.check_project_table(tables[PROJECT_TABLE_INDEX])
    #
    # def test_view_group_summary(self):
    #     response = self.factory_get(
    #         reverse(
    #             "forecast_group",
    #             kwargs={
    #                 'group_code': self.group.group_code
    #             },
    #         ),
    #         GroupView,
    #         group_code=self.group.group_code,
    #     )
    #
    #     self.assertEqual(response.status_code, 200)
    #     self.assertContains(response, "govuk-table")
    #     soup = BeautifulSoup(response.content, features="html.parser")
    #
    #     # Check that there are 4 tables on the page
    #     tables = soup.find_all("table", class_="govuk-table")
    #     assert len(tables) == 4
    #
    #     # Check that the first table displays the cost centre code
    #
    #     # Check that all the subtotal hierachy_rows exist
    #     table_rows = soup.find_all("tr", class_="govuk-table__row")
    #     assert len(table_rows) == 18
    #
    #     self.check_negative_value_formatted(soup)
    #
    #     self.check_hierarchy_table(tables[HIERARCHY_TABLE_INDEX],
    #                                self.directorate.directorate_name)
    #     # Check that the second table displays the programme and the correct totals
    #     self.check_programme_table(tables[PROGRAMME_TABLE_INDEX])
    #
    #     # Check that the third table displays the expenditure and the correct totals
    #     self.check_expenditure_table(tables[EXPENDITURE_TABLE_INDEX])
    #
    #     # Check that the second table displays the project and the correct totals
    #     self.check_project_table(tables[PROJECT_TABLE_INDEX])
    #
    # def test_view_dit_summary(self):
    #     response = self.factory_get(
    #         reverse("forecast_dit"),
    #         DITView,
    #     )
    #
    #     self.assertEqual(response.status_code, 200)
    #     self.assertContains(response, "govuk-table")
    #     soup = BeautifulSoup(response.content, features="html.parser")
    #
    #     # Check that there are 4 tables on the page
    #     tables = soup.find_all("table", class_="govuk-table")
    #     assert len(tables) == 4
    #
    #     # Check that the first table displays the cost centre code
    #
    #     # Check that all the subtotal hierarchy_rows exist
    #     table_rows = soup.find_all("tr", class_="govuk-table__row")
    #     assert len(table_rows) == 18
    #
    #     self.check_negative_value_formatted(soup)
    #
    #     self.check_hierarchy_table(tables[HIERARCHY_TABLE_INDEX],
    #                                self.group_name)
    #     # Check that the second table displays the programme and the correct totals
    #     self.check_programme_table(tables[PROGRAMME_TABLE_INDEX])
    #
    #     # Check that the third table displays the expenditure and the correct totals
    #     self.check_expenditure_table(tables[EXPENDITURE_TABLE_INDEX])
    #
    #     # Check that the second table displays the project and the correct totals
    #     self.check_project_table(tables[PROJECT_TABLE_INDEX])
