import io

from django.contrib.auth.models import Permission
from django.test import TestCase
from django.urls import reverse

from openpyxl import load_workbook

from chartofaccountDIT.test.factories import (
    HistoricalAnalysis1Factory,
    HistoricalAnalysis2Factory,
    HistoricalNaturalCodeFactory,
    HistoricalProgrammeCodeFactory,
    HistoricalProjectCodeFactory,
)

from core.models import FinancialYear
from core.test.test_base import RequestFactoryBase

from costcentre.test.factories import ArchivedCostCentreFactory

from forecast.views.view_forecast.export_forecast_data import (
    export_forecast_data_cost_centre,
    export_forecast_data_directorate,
    export_forecast_data_dit,
    # export_forecast_data_expenditure_detail_cost_centre,
    # export_forecast_data_expenditure_detail_directorate,
    # export_forecast_data_expenditure_detail_group,
    # export_forecast_data_expenditure_dit,
    export_forecast_data_group,
    export_forecast_data_programme_detail_directorate,
    export_forecast_data_programme_detail_dit,
    export_forecast_data_programme_detail_group,
    export_forecast_data_project_detail_cost_centre,
    export_forecast_data_project_detail_directorate,
    export_forecast_data_project_detail_dit,
    export_forecast_data_project_detail_group,
)

from previous_years.models import (
    ArchivedFinancialCode,
    ArchivedForecastData,
)


class DownloadPastYearForecastTest(TestCase, RequestFactoryBase):

    #     self.budget = create_budget(financial_code_obj, year_obj)
    #     self.year_total = self.amount_apr + self.amount_may
    #     self.underspend_total = self.budget - self.amount_apr - self.amount_may
    #     self.spend_to_date_total = self.amount_apr

    def setUp(self):
        RequestFactoryBase.__init__(self)
        # 2019 is created when the database is created, so it exists
        self.archived_year = 2019
        archived_year_obj = FinancialYear.objects.get(pk=self.archived_year)
        self.cost_centre_code = "109189"
        self.group_code = "1090TT"
        self.directorate_code = "10900T"
        self.natural_account_code = 52191003
        self.programme_code = "310940"
        self.project_code = "0123"
        self.analisys1 = "00798"
        self.analisys2 = "00321"
        cc_obj = ArchivedCostCentreFactory.create(
            cost_centre_code=self.cost_centre_code,
            directorate_code=self.directorate_code,
            group_code=self.group_code,
            financial_year=archived_year_obj,
        )
        project_obj = HistoricalProjectCodeFactory.create(
            project_code=self.project_code, financial_year=archived_year_obj
        )
        self.budget_type = "AME"
        programme_obj = HistoricalProgrammeCodeFactory.create(
            programme_code=self.programme_code,
            budget_type_id=self.budget_type,
            financial_year=archived_year_obj
        )
        nac_obj = HistoricalNaturalCodeFactory.create(
            natural_account_code=self.natural_account_code,
            economic_budget_code="CAPITAL",
            financial_year=archived_year_obj,
        )
        analysis2_obj = HistoricalAnalysis2Factory.create(
            analysis2_code=self.analisys2, financial_year=archived_year_obj
        )
        analysis1_obj = HistoricalAnalysis1Factory.create(
            analysis1_code=self.analisys1, financial_year=archived_year_obj
        )
        financial_code_obj = ArchivedFinancialCode.objects.create(
            programme=programme_obj,
            cost_centre=cc_obj,
            natural_account_code=nac_obj,
            analysis1_code=analysis1_obj,
            analysis2_code=analysis2_obj,
            project_code=project_obj,
            financial_year=archived_year_obj,
        )
        # __forecast_expenditure_type_name"
        self.expenditure_type_name = financial_code_obj.forecast_expenditure_type
        print(f'{self.expenditure_type_name }')
        previous_year_obj = ArchivedForecastData.objects.create(
            financial_year=archived_year_obj, financial_code=financial_code_obj,
        )
        self.outturn = {
            "budget": 1234500,
            "apr": 1200000,
            "may": 3412000,
            "jun": 9876000,
            "jul": 5468970,
            "aug": 3421900,
            "sep": 6901110,
            "oct": 7622200,
            "nov": 6955501,
            "dec": 8434422,
            "jan": 5264091,
            "feb": 4521111,
            "mar": 9090111,
            "adj01": 5464644,
            "adj02": 2118976,
            "adj03": 3135450,
        }
        previous_year_obj.budget = self.outturn["budget"] * 100
        previous_year_obj.apr = self.outturn["apr"] * 100
        previous_year_obj.may = self.outturn["may"] * 100
        previous_year_obj.jun = self.outturn["jun"] * 100
        previous_year_obj.jul = self.outturn["jul"] * 100
        previous_year_obj.aug = self.outturn["aug"] * 100
        previous_year_obj.sep = self.outturn["sep"] * 100
        previous_year_obj.oct = self.outturn["oct"] * 100
        previous_year_obj.nov = self.outturn["nov"] * 100
        previous_year_obj.dec = self.outturn["dec"] * 100
        previous_year_obj.jan = self.outturn["jan"] * 100
        previous_year_obj.feb = self.outturn["feb"] * 100
        previous_year_obj.mar = self.outturn["mar"] * 100
        previous_year_obj.adj1 = self.outturn["adj01"] * 100
        previous_year_obj.adj2 = self.outturn["adj02"] * 100
        previous_year_obj.adj3 = self.outturn["adj03"] * 100
        previous_year_obj.save()
        # Assign forecast view permission
        can_view_forecasts = Permission.objects.get(codename="can_view_forecasts")
        self.test_user.user_permissions.add(can_view_forecasts)
        self.test_user.save()

    def check_response_content(self, content):
        file = io.BytesIO(content)
        wb = load_workbook(filename=file)
        ws = wb.active
        # Check group
        assert ws["B1"].value == "Group code"
        assert ws["B2"].value == self.group_code
        assert ws["D1"].value == "Directorate code"
        assert ws["D2"].value == self.directorate_code
        assert ws["F1"].value == "Cost Centre code"
        assert ws["F2"].value == self.cost_centre_code

        assert ws["N1"].value == "PO/Actual NAC"
        assert ws["N2"].value == self.natural_account_code
        assert ws["Q1"].value == "Programme code"
        assert ws["Q2"].value == self.programme_code
        assert ws["W1"].value == "Project code"
        assert ws["W2"].value == self.project_code
        assert ws["S1"].value == "Contract code"
        assert ws["S2"].value == self.analisys1
        assert ws["U1"].value == "Market code"
        assert ws["U2"].value == self.analisys2
        assert ws["J1"].value == "Budget Type"
        assert ws["J2"].value == self.budget_type

        # print(f'{ws["G2"].value} {ws["H2"].value} {ws["J2"].value}')
        # check the figures
        assert ws["Y2"].value == self.outturn["budget"]
        assert ws["Z2"].value == self.outturn["apr"]
        assert ws["AA2"].value == self.outturn["may"]
        assert ws["AB2"].value == self.outturn["jun"]
        assert ws["AC2"].value == self.outturn["jul"]
        assert ws["AD2"].value == self.outturn["aug"]
        assert ws["AE2"].value == self.outturn["sep"]
        assert ws["AF2"].value == self.outturn["oct"]
        assert ws["AG2"].value == self.outturn["nov"]
        assert ws["AH2"].value == self.outturn["dec"]
        assert ws["AI2"].value == self.outturn["jan"]
        assert ws["AJ2"].value == self.outturn["feb"]
        assert ws["AK2"].value == self.outturn["mar"]
        assert ws["AL2"].value == self.outturn["adj01"]
        assert ws["AM2"].value == self.outturn["adj02"]
        assert ws["AN2"].value == self.outturn["adj03"]

    def test_dit_download(self):
        response = self.factory_get(
            reverse("export_forecast_data_dit", kwargs={"period": self.archived_year}),
            export_forecast_data_dit,
            period=self.archived_year,
        )
        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_group_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_group",
                kwargs={"group_code": self.group_code, "period": self.archived_year, },
            ),
            export_forecast_data_group,
            group_code=self.group_code,
            period=self.archived_year,
        )
        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_directorate_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_directorate",
                kwargs={
                    "directorate_code": self.directorate_code,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_directorate,
            directorate_code=self.directorate_code,
            period=self.archived_year,
        )
        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_cost_centre_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_cost_centre",
                kwargs={
                    "cost_centre": self.cost_centre_code,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_cost_centre,
            cost_centre=self.cost_centre_code,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_dit_project_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_dit",
                kwargs={
                    "project_code_id": self.project_code,
                    "period": self.archived_year
                },
            ),
            export_forecast_data_project_detail_dit,
            project_code_id=self.project_code,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)

        self.check_response_content(response.content)

    def test_group_project_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_group",
                kwargs={
                    "group_code": self.group_code,
                    "project_code_id": self.project_code,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_project_detail_group,
            group_code=self.group_code,
            project_code_id=self.project_code,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)

        self.check_response_content(response.content)

    def test_directorate_project_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_directorate",
                kwargs={
                    "directorate_code": self.directorate_code,
                    "project_code_id": self.project_code,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_project_detail_directorate,
            directorate_code=self.directorate_code,
            project_code_id=self.project_code,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)

    def test_cost_centre_project_download(self):
        response = self.factory_get(
            reverse(
                "export_forecast_data_project_detail_cost_centre",
                kwargs={
                    "cost_centre": self.cost_centre_code,
                    "project_code_id": self.project_code,
                    "period": self.archived_year,
                },
            ),
            export_forecast_data_project_detail_cost_centre,
            cost_centre=self.cost_centre_code,
            project_code_id=self.project_code,
            period=self.archived_year,
        )

        self.assertEqual(response.status_code, 200)
        self.check_response_content(response.content)


def test_directorate_programme_download(self):
    response = self.factory_get(
        reverse(
            "export_forecast_data_programme_detail_directorate",
            kwargs={
                "directorate_code": self.directorate_code,
                "programme_code_id": self.project_code,
                "forecast_expenditure_type_name": 1,
                "period": self.archived_year,
            },
        ),
        export_forecast_data_programme_detail_directorate,
        directorate_code=self.directorate_code,
        programme_code_id=self.programme_code,
        forecast_expenditure_type_name=self.project_code,
        period=self.archived_year,
    )

    self.assertEqual(response.status_code, 200)
    self.check_response_content(response.content)
