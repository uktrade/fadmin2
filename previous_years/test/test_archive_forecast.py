import os

from django.core.management import call_command
from django.test import (
    RequestFactory,
    TestCase,
)

from openpyxl import Workbook

from chartofaccountDIT.test.factories import (
    HistoricalAnalysis1Factory,
    HistoricalAnalysis2Factory,
    HistoricalNaturalCodeFactory,
    HistoricalProgrammeCodeFactory,
    HistoricalProjectCodeFactory,
)

from core.models import FinancialYear
from core.test.test_base import RequestFactoryBase

from costcentre.test.factories import ArchivedCostCentreFactory

from previous_years.import_previous_year import (
    ANALYSIS2_HEADER,
    ANALYSIS_HEADER,
    ArchiveYearError,
    COST_CENTRE_HEADER,
    DATA_HEADERS,
    NAC_HEADER,
    PROGRAMME_HEADER,
    PROJECT_HEADER,
    VALID_WS_NAME,
    upload_previous_year,
)
from previous_years.models import (
    ArchivedFinancialCode,
    ArchivedForecastData,
)


from upload_file.models import FileUpload


class ImportPreviousYearForecastTest(TestCase, RequestFactoryBase):
    def setUp(self):
        RequestFactoryBase.__init__(self)
        self.factory = RequestFactory()
        # 2019 is created when the database is created
        self.archived_year = 2019
        self.archived_year_obj = FinancialYear.objects.get(pk=self.archived_year)
        self.cost_centre_code = "109189"
        self.natural_account_code = 52191003
        self.programme_code = "310940"
        self.project_code = "0123"
        self.analisys1 = "00798"
        self.analisys2 = "00321"
        ArchivedCostCentreFactory.create(
            cost_centre_code=self.cost_centre_code,
            financial_year=self.archived_year_obj,
        )
        HistoricalProjectCodeFactory.create(
            project_code=self.project_code, financial_year=self.archived_year_obj
        )
        HistoricalProgrammeCodeFactory.create(
            programme_code=self.programme_code, financial_year=self.archived_year_obj
        )
        HistoricalNaturalCodeFactory.create(
            natural_account_code=self.natural_account_code,
            economic_budget_code="CAPITAL",
            financial_year=self.archived_year_obj,
        )
        HistoricalAnalysis2Factory.create(
            analysis2_code=self.analisys2, financial_year=self.archived_year_obj
        )
        HistoricalAnalysis1Factory.create(
            analysis1_code=self.analisys1, financial_year=self.archived_year_obj
        )

    def tearDown(self):
        if os.path.exists(self.excel_file_name):
            os.remove(self.excel_file_name)

    def create_workbook(self):
        wb = Workbook()
        self.data_worksheet = wb.active
        self.data_worksheet.title = VALID_WS_NAME
        col_index = 1
        self.data_worksheet.cell(column=col_index, row=1, value=COST_CENTRE_HEADER)
        self.data_worksheet.cell(column=col_index, row=2, value=self.cost_centre_code)
        col_index += 1
        self.data_worksheet.cell(column=col_index, row=1, value=NAC_HEADER)
        self.data_worksheet.cell(
            column=col_index, row=2, value=self.natural_account_code
        )
        col_index += 1
        self.data_worksheet.cell(column=col_index, row=1, value=PROGRAMME_HEADER)
        self.data_worksheet.cell(column=col_index, row=2, value=self.programme_code)
        col_index += 1
        self.data_worksheet.cell(column=col_index, row=1, value=PROJECT_HEADER)
        self.data_worksheet.cell(column=col_index, row=2, value=self.project_code)
        col_index += 1
        self.data_worksheet.cell(column=col_index, row=1, value=ANALYSIS_HEADER)
        self.data_worksheet.cell(column=col_index, row=2, value=self.analisys1)
        col_index += 1
        self.data_worksheet.cell(column=col_index, row=1, value=ANALYSIS2_HEADER)
        self.data_worksheet.cell(column=col_index, row=2, value=self.analisys2)

        for month in DATA_HEADERS:
            col_index += 1
            self.data_worksheet.cell(column=col_index, row=1, value=month)
        self.data_worksheet.cell(column=col_index, row=2, value=col_index * 13)
        self.data_worksheet.cell(column=col_index, row=3, value=col_index * 7)
        self.excel_file_name = "dummy.xlsx"
        wb.save(filename=self.excel_file_name)

    def test_upload_previous_year(self):
        self.assertEqual(ArchivedFinancialCode.objects.all().count(), 0)
        self.assertEqual(ArchivedForecastData.objects.all().count(), 0)
        self.create_workbook()
        file_upload_obj = FileUpload(
            document_file_name=self.excel_file_name,
            document_type=FileUpload.PREVIOUSYEAR,
            file_location=FileUpload.LOCALFILE,
        )
        file_upload_obj.save()

        upload_previous_year(
            self.data_worksheet, self.archived_year, file_upload_obj,
        )
        self.assertEqual(ArchivedFinancialCode.objects.all().count(), 1)
        self.assertEqual(ArchivedForecastData.objects.all().count(), 1)

    def test_upload_wrong(self):
        self.create_workbook()
        file_upload_obj = FileUpload(
            document_file_name=self.excel_file_name,
            document_type=FileUpload.PREVIOUSYEAR,
            file_location=FileUpload.LOCALFILE,
        )
        file_upload_obj.save()
        with self.assertRaises(ArchiveYearError):
            upload_previous_year(
                self.data_worksheet, self.archived_year + 1, file_upload_obj,
            )

    def test_command(self):
        self.assertEqual(ArchivedFinancialCode.objects.all().count(), 0)
        self.assertEqual(ArchivedForecastData.objects.all().count(), 0)
        self.create_workbook()
        call_command(
            "upload_previous_year", "dummy.xlsx", self.archived_year,
        )
        self.assertEqual(ArchivedFinancialCode.objects.all().count(), 1)
        self.assertEqual(ArchivedForecastData.objects.all().count(), 1)


# TODO
# test for non numeric value
# check the totals
